"""Common admin functions."""
import csv
import time
from django.http import HttpResponse
from django.contrib import messages


def dump_to_csv(modeladmin, request, qs):
    """
    These takes in a Django queryset and spits out a CSV file.

    Generic method for any queryset
    """
    model = qs.model
    file_id = 'System_%s_%d' % (model.__name__, int(time.time()))
    file_name = 'attachment; filename=%s.csv' % (file_id)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = file_name
    writer = csv.writer(response, csv.excel)

    headers = []
    for field in model._meta.fields:
        headers.append(field.name)
    writer.writerow(headers)

    for obj in qs:
        row = []
        for field in headers:
            val = getattr(obj, field)
            if callable(val):
                val = val()
            if isinstance(val, str):
                # val.encode("utf-8")
                val = str(val)
            row.append(val)
        writer.writerow(row)
    return response


dump_to_csv.short_description = "Dump to CSV"


def export_xls(modeladmin, request, queryset):
    """Method to export as excel."""
    import xlwt

    model = queryset.model
    file_id = 'System_%s_%d' % (model.__name__, int(time.time()))
    file_name = 'attachment; filename=%s.xls' % (file_id)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = file_name
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("Items List")
    row_num = 0

    headers = []
    for field in model._meta.fields:
        headers.append(field.name)

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in range(len(headers)):
        ws.write(row_num, col_num, headers[col_num], font_style)

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1
    for obj in queryset:
        row_num += 1
        row = []
        for field in headers:
            val = getattr(obj, field)
            if callable(val):
                val = val()
            if isinstance(val, str):
                # val.encode("utf-8")
                val = str(val)
            row.append(val)
        for col_num in range(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
    wb.save(response)
    return response


export_xls.short_description = "Export XLS"


def export_xlsx(modeladmin, request, queryset):
    """Export as xlsx."""
    import openpyxl

    model = queryset.model
    file_id = 'System_%s_%d' % (model.__name__, int(time.time()))
    file_name = 'attachment; filename=%s.xlsx' % (file_id)
    fmt = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response = HttpResponse(content_type=fmt)
    response['Content-Disposition'] = file_name
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Items List"

    headers = []
    for field in model._meta.fields:
        headers.append(field.name)

    row_num = 0
    for col_num in range(len(headers)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = headers[col_num]

    for obj in queryset:
        row_num += 1
        row = []
        for field in headers:
            val = getattr(obj, field)
            if callable(val):
                val = val()
            if isinstance(val, str):
                # val.encode("utf-8")
                val = str(val)
            row.append(val)
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            # c.style.alignment.wrap_text = True

    wb.save(response)
    return response


export_xlsx.short_description = "Export XLSX"


def void_records(modeladmin, request, queryset):
    """
    These takes the queryset and sets the records value

    for is_void to True
    """
    updated = queryset.update(is_void=True)
    message = '%d Records Successfully voided (Soft delete).' % (updated)
    messages.info(request, message)


void_records.short_description = "Void Records"
